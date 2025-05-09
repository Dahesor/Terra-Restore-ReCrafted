import openpyxl
import json
import os

def generate_lang_json(xlsx_path, output_dir=".",types=["generic"]):
    # 加载工作簿，data_only=True 以获取公式结果
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    for type in types:
        w_this = wb[type]
        result = read_this_excel(w_this)
        save_json_file(result, type, output_dir)
    return

def make_id(en_prefix):
    return "prefix." + en_prefix.strip().lower().replace(" ", "_").replace("-", "_")

def save_json_file(data, type, output_dir):
    zh_path = os.path.join(output_dir, f"{type}_prefix", "lang", "zh_cn.json")
    en_path = os.path.join(output_dir, f"{type}_prefix", "lang", "en_us.json")

    os.makedirs(os.path.dirname(zh_path), exist_ok=True)
    os.makedirs(os.path.dirname(en_path), exist_ok=True)

    current_working_dir = os.getcwd()
    print(current_working_dir)

    with open(zh_path, "w", encoding="utf-8") as f:
        json.dump(data[0], f, ensure_ascii=False, indent=2)
    with open(en_path, "w", encoding="utf-8") as f:
        json.dump(data[1], f, ensure_ascii=False, indent=2)
    print(f"✅ 已输出语言文件：{type}")

def read_this_excel(ws):
    headers = [cell.value.strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    try:
        zh_index = headers.index("中文名")
        en_index = headers.index("英文前缀")
    except ValueError:
        raise Exception("请确保 Excel 中包含列：中文名、英文前缀")

    lang_zh = {}
    lang_en = {}

    for row in ws.iter_rows(min_row=2):
        zh = row[zh_index].value
        en = row[en_index].value
        if not zh or not en:
            continue

        entry_id = make_id(str(en))
        lang_zh[entry_id] = str(zh)
        lang_en[entry_id] = str(en).capitalize()
    return lang_zh, lang_en
