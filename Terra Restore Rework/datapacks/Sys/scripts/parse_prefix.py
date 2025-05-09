import openpyxl
import json
import os

def make_id(en_prefix):
    return "prefix." + en_prefix.strip().lower().replace(" ", "_").replace("-", "_")

def generate_lang_json(xlsx_path, output_dir="../../../../resourcepacks/Resource Pack for Terra Restore/lang/assets/armo_prefix/lang"):
    # 加载工作簿，data_only=True 以获取公式结果
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 获取表头
    headers = [cell.value.strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    try:
        zh_index = headers.index("中文名")
        en_index = headers.index("英文前缀")
    except ValueError:
        raise Exception("请确保 Excel 中包含列：中文名、英文前缀")

    lang_zh = {}
    lang_en = {}

    for row in ws.iter_rows(min_row=2):
        zh = row[zh_index].value
        en = row[en_index].value
        if not zh or not en:
            continue

        entry_id = make_id(str(en))
        lang_zh[entry_id] = str(zh)
        lang_en[entry_id] = str(en).capitalize()

    # 保存为 JSON
    zh_path = output_dir + "/zh_cn.json"
    en_path = output_dir + "/en_us.json"

    with open(zh_path, "w", encoding="utf-8") as f:
        json.dump(lang_zh, f, ensure_ascii=False, indent=2)
    with open(en_path, "w", encoding="utf-8") as f:
        json.dump(lang_en, f, ensure_ascii=False, indent=2)

    print(f"✅ 已输出语言文件：\n中文: {zh_path}\n英文: {en_path}")

def safe_int(val):
    try:
        return int(val)
    except:
        return 0

def generate_mcfunction(xlsx_path, output_path="data/registry/function/prefix/armor/package.mcfunction"):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(wb.sheetnames)
    ws = wb["工作表1"]

    # 读取表头
    headers = [cell.value.strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # 找出各列索引
    col_map = {key: headers.index(key) for key in ["中文名", "英文前缀", "评分", "稀有度"]}
    attributes = ["STR", "CON", "DEX", "POW", "EDU", "INT", "crit_chance", "crit_damage"]
    for attr in attributes:
        if attr not in headers:
            headers.append(attr)

    attr_indices = {attr: headers.index(attr) for attr in attributes}

    # 生成 .mcfunction 行
    lines = []
    for row in ws.iter_rows(min_row=2):
        zh_name = row[col_map["中文名"]].value
        en_prefix = row[col_map["英文前缀"]].value
        score = safe_int(row[col_map["评分"]].value)
        rarity = safe_int(row[col_map["稀有度"]].value) if row[col_map["稀有度"]].value else 0

        prefix_id = "" + en_prefix.strip().lower().replace(" ", "_").replace("-", "_")
        negative = "1b" if rarity<0 else "0b"

        # 构建属性 scale
        stats = []
        for attr in attributes:
            val = row[attr_indices[attr]].value
            if val is None:
                continue
            val = safe_int(val)
            if val != 0:
                if attr in ["crit_chance", "crit_damage"]:
                    add = int(val * 10)
                    stats.append(f"{attr}:{{add:{add}}}")
                else:
                    scale = round(val / 100, 3)
                    stats.append(f"{attr}:{{scale:{scale}d}}")

        stats_effects = "{" + ",".join(stats) + "}"

        # 添加注释与语句
        lines.append(f"# {zh_name}")
        lines.append(
            f'data modify storage registry:prefix new append value '
            f'{{id:"{prefix_id}",for:["helmet","chestplate","leggings","boots"],'
            f'negative:{negative},rarity:{rarity},stats_effects:{stats_effects},att_effects:{{}},value:{score}}}'
        )
        lines.append("")

    # 写入文件
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"✅ .mcfunction 文件已生成：{output_path}")

if __name__ == "__main__":
    source = "scripts/大地复苏_prefix.xlsx"
    generate_mcfunction(source)
    generate_lang_json(source)
